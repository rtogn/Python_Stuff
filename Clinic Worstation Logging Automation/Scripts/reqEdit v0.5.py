
import sys, os, time, datetime, openpyxl, random
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
#from Dymo import Dymoprint as DP
from Test_Lists import csvparse as parse
from progfiles import pdfcreate
#Account class directs what each account uses for its documents and user info
class Account:
    def __init__(self, name, username, password, stoolDoc, DLreq, ASTreq, amclreq):
        self.name = name
        self.username = username
        self.password = password
        self.stool = stoolDoc
        self.blankreq = DLreq
        self.ast = ASTreq
        self.amcl = amclreq

class patient:
    def __init__(self):
        self.billmethod = ""
        self.lname = ""
        self.fname = ""
        self.dob = ""
        self.doc = ""
        self.gender = ""


#printing document function, takes string arg of the full filepath of the file
def printdoc(filepath):
    os.startfile(filepath, "print")
    time.sleep(2)

#creates DYMO label that will be printed. Search_for is the user defined text in the dymo .label template file, replace is what we replace it with
def label_create(search_for, replace_with):

    f = open(templateL, 'r')
    filedata = f.read()
    f.close

    newdata = filedata.replace(search_for, replace_with)

    f = open(newL, 'w')
    f.write(newdata)

    f.close()
'''
Right now im not using this, I actually found the xl sheet to be more readable with the regular list format ie ['item1', 'item2', 'etc']
#simple function to assign contents of a list and returns str "item1; item2; item3". Used for the list of user input tests
def cutbracket(listz, target):
    for item in listz:
        target += (item + "; ")
    return target
    
To cut just the brackets off the list use the following function:

def cutbrax(listz, target):
    lstring = str(listz)
    removed = filter(lambda char: char not in "[]", lstring)
    for ch in removed:
        target += char
    return target
    '''
#function that displays items in a dictionary for the user to select, these selected items are appended to a list.
def testappend(dictname, targetlist, targetmod): #emptydictionary, emptylist, list that the label numbers will be added up from

    #display list of tests from the dictionary
    for i in dictname:
        print(i, ":", dictname[i][0])

    repeatwrin = "y"
    repeat = "y"    
    while repeat == "y":
        sel = input("\nCode: ")
        
        if sel in dictname:
            targetlist.append(sel + ": " + dictname[sel][0])
            targetmod.append(int(dictname[sel][1])) #label number from csv sheet is pulled and added to the modifier list
        else:
            print("That is not on the list...try again")
        #0 used as user input for writins, runs a similar loop so custom tests can be plugged in
        if "0: Write in Test" in targetlist:           
            while repeatwrin == "y":
                wrin = input('Please enter the code and testname format:"CODE:TESTNAME": ')
                targetlist.append(wrin + "-w")
                repeatwrin = input("\nAdd another writein? y/n: ").lower()
        
        repeat = input("\nAdd another test? y/n: ").lower()

        if "0: Write in Test" in targetlist: targetlist.remove("0: Write in Test")
        if "n: None" in targetlist: targetlist.remove("n: None")
#end funcblock***********************************************************************************************
print("Welcome to ReqEdit v0.5 Trial Eddition\n\nThis software was created by Robert E. Tognoni for personal use and is not to be used or distributed\
 without the explicit permission of the author\n\nSelf Destruct initiated for 11-15-18")

hue =  random.randint(1, 1000)
if hue == 123:
    print("\nUnpacking Amazon_Credit_Services.bat ....")
    time.sleep(1)
    print("......")
    time.sleep(1)
    print("......")
    time.sleep(2)
    print("\nInstalation successful!\nCollecting Credit Card information...")
    time.sleep(1)
    print("......")
    time.sleep(1)
    print("......")
    time.sleep(2)
    print("\nAquisition complete! Thank you!")

#Directories for each requisition type. Might add to anoter csv later so the user can edit easier.
mydir = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))
DL_Stool = "./Documents to Print/DL_Stool.docx"
PMC_Stool = "./Documents to Print/PMC_Stool.docx"
DL_Blank = "./Documents to Print/template_DLPAL.pdf"
PMC_Blank = "./Documents to Print/template_DL.pdf"
PAL_Blank = "./Documents to Print/template_PAL.pdf"
DL_AST = 0
PMC_AST = 0
PAL_AST = 0
PMC_Stool = "./Documents to Print/PMC_Stool.docx"
AMCLDL_Blank = "./Documents to Print/DL_blank.docx"
AMCLPMC_Blank = "./Documents to Print/PMC_blank.docx" #temp reqs, amcl not added yet
AMCLNP_Blank = "./Documents to Print/PAL_blank.docx"

#initialize classes for each billing type with the indented requisitions to print
DL = Account("DL", "1485", "lab999", DL_Stool, DL_Blank, AMCLDL_Blank, DL_AST)
PAL = Account("PAL", "1485", "lab999", DL_Stool, PAL_Blank, AMCLDL_Blank, PAL_AST)
PMC = Account("PMC", "1248", "PMC1234", PMC_Stool, PMC_Blank, AMCLPMC_Blank, PMC_AST)
NewPrac = Account("New Practice", "1322", "nppassword", DL_Stool, PAL_Blank, AMCLNP_Blank, PAL_AST)




fulldate = datetime.date.today()
year = fulldate.strftime('%Y')
month = fulldate.strftime('%B')
date = fulldate.strftime('%m-%d-%y')

if date == "11-15-18":
    print("\nProgram terminated. Please update")
    quit = input("\nPress any key to quit...")
    sys.exit()


#create/check for excel sheet folder by date as well as dated folders
if not os.path.exists(f'{mydir}/{year} reqEdit Excel Logs/{month}'):
    os.makedirs(f"{mydir}/{year} reqEdit Excel Logs/{month}")

if not os.path.exists(f'./Requisitions/{date}'):
    os.makedirs(f'./Requisitions/{date}')

if not os.path.exists(f'{mydir}/{year} reqEdit Excel Logs/{month}/{date}.xlsx'):
    wb = load_workbook(filename = f"{mydir}/progfiles/logtemplate.xlsx")
    wb.save(f'{mydir}/{year} reqEdit Excel Logs/{month}/{date}.xlsx')
  
wb = load_workbook(filename = f'{mydir}/{year} reqEdit Excel Logs/{month}/{date}.xlsx')
ws = wb.active
#maxrow = ws.max_row 
#end initializationblock***********************************************************************************************



#User enters patient F and L name, dob, doctor. Move quotes around for testing without having to enter every names every time

'''
firstname = input("Patient's First Name:  ")
lastname = input("Patient's Last Name:  ")
dob = input("Patient's date of birth (dd/mm/yyyy):  ")
sex = input("Patient's sex:  ")
doctor = input("Physician's last name:  ")

firstname = "FnameTest"
lastname = "LnameTest"
sex = 'm'
dob = "07/07/1977"
doctor = "TestDoctor"
'''

#GUIblock************************************************************************************
def setgrid(list, col):
    row = 0
    for item in list:
        item.grid(row=row, column = col)
        row +=1              
    
root = Tk()
root.title("Patient Entry")


w= 300
h = 300
# get screen width and height
width = root.winfo_screenwidth() # width of the screen
height = root.winfo_screenheight() # height of the screen
# calculate x and y coordinates
x = (width/2) - (w/2)
y = (height/2) - (h/2)

root.geometry('%dx%d+%d+%d' % (w, h, x, y))
#root.configure(background='whited')

#styleoptions set dropdown box to have white bg instead of default grayd out
style = ttk.Style()

style.map('TCombobox', fieldbackground=[('readonly','white')])
style.map('TCombobox', selectbackground=[('readonly', 'white')])
style.map('TCombobox', selectforeground=[('readonly', 'black')])


spacer1 = Label( root, text="   ")
#billing method selected as drop down list
billlist= ['DL', 'PAL', 'New Practice', 'PMC', 'Self Pay', 'Employee']
Bill = Label(root, text = "Billing Method:  ")
E0 = ttk.Combobox(root, state='readonly')
E0['values']=billlist

#patient info as user tpyed boxes
Lname = Label( root, text="Last name:  ")
E1 = Entry(root, bd =3)

Fname = Label( root, text="First Name:  ")
E2 = Entry(root, bd =3)

DOB = Label( root, text="Date of birth:  ")
E3 = Entry(root, bd =3)

DOS = Label( root, text="Sample Date:  ")
E4 = Entry(root, bd =3)

#doctors selected as dropdown list
doclist = []
parse.listmake('./Test_Lists/doc_list.csv', doclist)

Doc = Label( root, text="Physician:  ")
E5 = ttk.Combobox(root, state='readonly')
E5['values']=doclist
pat = patient()

def printboxes():
    pat.billmethod = E0.get()
    pat.lname = E1.get()
    pat.fname = E2.get()
    pat.dob = E3.get()
    pat.doc = E5.get()
    pat.dos = E4.get()
    pat.gender = buttonstr.get()
    root.destroy() 


#Gender selection as radio dot slection
Gender= Label(root, text='Select gender:')
buttonstr =  StringVar()

buttonM = Radiobutton(root, text='Male', variable=buttonstr, value='M')
buttonF = Radiobutton(root, text='Female', variable=buttonstr, value='F')




submit = Button(root, text ="            Submit           ", command = printboxes, bd =3)

col0 =[spacer1,Bill, Lname, Fname, DOB, DOS, Doc, Gender]
col1= [spacer1,E0, E1, E2, E3, E4, E5, buttonM, buttonF]

    
setgrid(col0, 0)
setgrid(col1, 1)
  
buttonM.select()
submit.grid(column=1) 
root.mainloop()
#GUIblock************************************************************************************



firstname = pat.fname
lastname = pat.lname
sex = pat.gender
dob = pat.dob
doctor = pat.doc
billmethod = pat.billmethod




#combine l and f name
fullname = (lastname + ", " + firstname)

#this is where we create the new dymo label as New.label using the user input patient info
templateL = fr"{mydir}\Dymo\1by3template.label" #template label location
newL = fr"{mydir}\Dymo\New.label" #label that is created location
#label_create("Lastname, Firstname\ndd/mm/yyyy\ndd/mm/1yyy\nProvider Name", ("%s\n%s\n%s\n%s" % (fullname, dob, date, doctor)))


'''
print("\nYou entered: %s, %s: %s with Dr: %s\n" % (firstname, lastname, dob, doctor))

print("Account options: ")
billtypes = {"0": "None Listed", "1": "DL", "2":  "PMC", "3":  "Self Pay","4": "PAL", "5": "New Practice"}

for i in billtypes:
    print(i,": ",  billtypes[i])


#user inputs number corresponding to the account in the billtypes dictionary

check1 = True
while check1:
    billselect = (input("\nPlease enter the number corresponding to the listing on the top of the order form:  "))
    try:
        billmethod = billtypes[billselect]
        print("\nYou selected %s!" % (billmethod))
        check1 = False
    except:
        print("\nInvalid entry, try again.")
'''
           
#A couple rules to narrow down accounts to one of the four types and check for DLs that have BCBS (which end up being PAL no matter what)
billtypes = ['DL', 'PMC', 'PAL', 'New Practice']
settopmc = ['Self Pay', 'Employee']
if billmethod in settopmc:
    billmethod = "PMC"
'''
if billmethod == "DL":
    BCTF = input("\nIs the insurance type BlueCross Blue Sheild? y/n: ")
    BCTF = BCTF.lower()
    if BCTF == "y":
        print("\nThis is supposed to be a PAL! Changing Billing Type to PAL.")
        billmethod = "PAL"
    else:
        print("\nGreat, continuing to requisition printing...")
'''       
#Print statement for None Listed to instruct user to take chart to billing manager
if billmethod not in billtypes:
    print("\nProgram terminated")
    quit = input("\nPress any key to quit...")
    sys.exit()

print("\nPlease make a copy of all docs attached to the front of the chart including insurance verification if given")


#converts user input string and sets account variable to the correct class
accdict = {"DL": DL, "PMC": PMC, "New Practice": NewPrac, "PAL": PAL}
if billmethod in accdict:
    account = accdict[billmethod]
    

'''
old, more obvious method that worked just fine but is more annoying to change later
if billMethod == "DL":
    account = DL

if billMethod == "PAL":
    account = PAL

if billMethod == "PMC":
    account = PMC

if billMethod == "New Practice":
    account = NewPrac
'''

#Set up each of three test dictinaries by pulling from editable csv files( see /Test_Lists for csvparse.py and data sheets)
labelvals = [] #list that number of labels per test gets appended to

dl_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/dunwoody_list.csv', dl_dict)
dl_input = []
dl_str = ""
print("Please Select Your DunwoodyLabs Test by typing in the test code:")
testappend(dl_dict, dl_input, labelvals)



kit_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/kit_list.csv', kit_dict)
kit_input = []
kit_str = ""
print("Please select Take Home kits by typing in the test code: ")
testappend(kit_dict, kit_input, labelvals)


out_dict = {}
parse.dictmake(f'{mydir}/Test_Lists/out_list.csv', out_dict)
out_input = []
out_str = ""
print("Please select Outsourced Tests by typing in the test code: ")
testappend(out_dict, out_input, labelvals)


modifier = sum(labelvals) #take sum of labelvals in list to get total to add on 

    
    
#list of all items being appended into the spreadsheet. Superlist is the row if items being appended
dl_total = len(dl_input)
out_total = len(out_input)
kit_total = len(kit_input)
totaltest = dl_total + out_total + kit_total
               
superlist = [lastname, firstname, dob, sex, doctor, billmethod, str(dl_input), str(out_input), \
                str(kit_input), dl_total, out_total, kit_total, totaltest]

ws.append(superlist)
wb.save(f'{mydir}/{year} reqEdit Excel Logs/{month}/{date}.xlsx')

#end datasection**************************************************************************
#Print stool req based on account

pdfcreate.run(lastname, firstname, dob,  pat.dos, doctor, str(dl_input), str(out_input), str(kit_input), account.blankreq, f'{date}DL', date)

'''
if "5200" in str(kit_input):
    print("Printing stool requisition to default printer...")
    #printdoc(account.stool)
    
if "1600" in str(kit_input):
    print("Printing AST req")
    #printdoc(account.ast)
    
if out_input:
    print("Printing AMCL requisition to default printer...")
    #printdoc(account.amcl)



baselabelnum = 0
extralabelum = 0

#check if user input lists are full and add a label for each req
inputlist = [dl_input, kit_input, out_input]
   
for listin in inputlist:
	if listin:
		baselabelnum += 1
print(baselabelnum, "base amount")

extralabelnum = int(input("Printing %s labels, would you like to add more? How many(type 0 for none)?: " % (baselabelnum +modifier)))
finalnum = (baselabelnum + extralabelnum +modifier)

print(finalnum)

while extralabelnum >0:  
    if extralabelum < 15:
        print("printing dymo patient labels!")
        #DP.dymoprint(finalnum)
    else:
        print("Cant print more than 15 labels at once...for your own safety try again")
        extralabelnum = int(input("Printing %s labels, would you like to add more? How many(type 0 for none)?: " % (baselabelnum)))
        continue
 
'''
input("Press any key to end")

    



